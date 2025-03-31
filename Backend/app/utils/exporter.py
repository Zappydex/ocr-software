import pandas as pd
import io
import logging
from typing import List
from app.models import Invoice
from app.config import settings
import asyncio
from concurrent.futures import ThreadPoolExecutor
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
import csv
from decimal import Decimal
import re

logger = logging.getLogger(__name__)

class InvoiceExporter:
    def __init__(self):
        self.columns = [
            "Filename", "Invoice Number", "Vendor Name", "Address", 
            "Invoice Date", "Grand Total", "Taxes", "Final Total", 
            "Description", "Pages"
        ]
        self.executor = ThreadPoolExecutor(max_workers=settings.MAX_WORKERS)
        self.default_currency = '$'  # Default currency symbol

    async def export_invoices(self, invoices: List[Invoice], format: str) -> io.BytesIO:
        try:
            df = await self._create_dataframe(invoices)
            if format.lower() == 'csv':
                return await self._export_to_csv(df)
            elif format.lower() == 'excel':
                return await self._export_to_excel(df)
            else:
                raise ValueError(f"Unsupported export format: {format}")
        except Exception as e:
            logger.error(f"Error during invoice export: {str(e)}")
            raise

    async def _create_dataframe(self, invoices: List[Invoice]) -> pd.DataFrame:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(self.executor, self._create_dataframe_sync, invoices)

    def _format_decimal(self, value):
        """Format decimal values intelligently based on their precision"""
        if value is None or value == 0:
            return None
            
        # Convert to string to check decimal places
        str_value = str(value)
        if '.' in str_value:
            integer_part, decimal_part = str_value.split('.')
            
            # If more than 4 decimal places, truncate to 4
            if len(decimal_part) > 4:
                return f"{float(value):.4f}"
            # If 1-2 decimal places, keep 2 decimal places
            elif len(decimal_part) <= 2:
                return f"{float(value):.2f}"
            # If 3-4 decimal places, keep as is
            else:
                return str_value
        else:
            return str_value

    def _extract_numeric_value(self, formatted_value):
        """Extract numeric value from formatted string with currency symbol"""
        if formatted_value is None:
            return 0
        
        if isinstance(formatted_value, (int, float, Decimal)):
            return float(formatted_value)
            
        # Remove currency symbol and convert to float
        numeric_str = re.sub(r'[^\d.]', '', str(formatted_value))
        try:
            return float(numeric_str) if numeric_str else 0
        except ValueError:
            return 0

    def _create_dataframe_sync(self, invoices: List[Invoice]) -> pd.DataFrame:
        data = []
        grand_total_sum = 0
        final_total_sum = 0
        
        for index, invoice in enumerate(invoices, 1):
            # Combine address fields into a single address
            address_parts = [
                invoice.vendor.address.street,
                invoice.vendor.address.city,
                invoice.vendor.address.state,
                invoice.vendor.address.postal_code,
                invoice.vendor.address.country
            ]
            address = ", ".join([part for part in address_parts if part])
            
            # Use "Purchase X" as the description
            description = f"Purchase {index}"
            
            # Format monetary values with currency symbol
            grand_total = None
            if invoice.grand_total is not None:
                grand_total_sum += float(invoice.grand_total)
                grand_total = f"{self.default_currency}{self._format_decimal(invoice.grand_total)}"
                
            taxes = None
            if invoice.taxes is not None:
                taxes = f"{self.default_currency}{self._format_decimal(invoice.taxes)}"
                
            final_total = None
            if invoice.final_total is not None:
                final_total_sum += float(invoice.final_total)
                final_total = f"{self.default_currency}{self._format_decimal(invoice.final_total)}"
            
            row = {
                "Filename": invoice.filename,
                "Invoice Number": invoice.invoice_number,
                "Vendor Name": invoice.vendor.name,
                "Address": address,
                "Invoice Date": invoice.invoice_date,
                "Grand Total": grand_total,
                "Taxes": taxes,
                "Final Total": final_total,
                "Description": description,
                "Pages": index 
            }
            data.append(row)

        # Create DataFrame
        df = pd.DataFrame(data, columns=self.columns)
        
        # Add Sum Total row
        sum_row = {col: "" for col in self.columns}
        sum_row["Vendor Name"] = "TOTAL"
        sum_row["Grand Total"] = f"{self.default_currency}{self._format_decimal(grand_total_sum)}"
        sum_row["Final Total"] = f"{self.default_currency}{self._format_decimal(final_total_sum)}"
        
        # Append sum row to DataFrame
        df = pd.concat([df, pd.DataFrame([sum_row], columns=self.columns)], ignore_index=True)
        
        return df

    async def _export_to_csv(self, df: pd.DataFrame) -> io.BytesIO:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(self.executor, self._export_to_csv_sync, df)

    def _export_to_csv_sync(self, df: pd.DataFrame) -> io.BytesIO:
        output = io.BytesIO()
        # Use original CSV export settings to maintain format
        df.to_csv(output, index=False)
        output.seek(0)
        return output

    async def _export_to_excel(self, df: pd.DataFrame) -> io.BytesIO:
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(self.executor, self._export_to_excel_sync, df)

    def _export_to_excel_sync(self, df: pd.DataFrame) -> io.BytesIO:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Invoices', index=False)
            
            workbook = writer.book
            sheet = workbook['Invoices']
            
            # Define medium border style for a more professional look
            medium_border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
            
            # Apply formatting to all cells
            for row_idx, row in enumerate(sheet.iter_rows(), 1):
                for cell in row:
                    cell.border = medium_border
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
                    
                    # Highlight the sum row
                    if row_idx == len(df) + 1:  # +1 for header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    
            # Make headers bold
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            
            # Auto-adjust column widths with extra padding for text-heavy columns
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                column_name = sheet[f"{column_letter}1"].value
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Add extra padding for text-heavy columns
                if column_name in ["Vendor Name", "Address", "Description"]:
                    adjusted_width = max_length + 5
                else:
                    adjusted_width = max_length + 2
                    
                sheet.column_dimensions[column_letter].width = adjusted_width

        output.seek(0)
        return output

async def export_invoices(invoices: List[Invoice], format: str) -> io.BytesIO:
    exporter = InvoiceExporter()
    return await exporter.export_invoices(invoices, format)
